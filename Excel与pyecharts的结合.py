from pyecharts import options as opts
from pyecharts.globals import ThemeType
from pyecharts.charts import Bar
from pyecharts.globals import ThemeType
from openpyxl import load_workbook

workbook=load_workbook(filename='first.xlsx')
sheet=workbook['Sheet1']
bar=Bar()
bar=Bar(init_opts=opts.InitOpts(theme=ThemeType.SHINE))
b=[]
b2=[]
for i1 in range(5):
    b.insert(i1,sheet.cell(row=1,column=i1+1).value)

for i2 in range(5):
    b2.insert(i2,sheet.cell(row=2,column=i2+1).value)
print(b)
bar.add_xaxis(b)
bar.add_yaxis('第一天',b2)

bar.set_global_opts(title_opts=opts.TitleOpts(
    title="第一天数据统计",
    subtitle="我是标题的弟弟"),
    xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(rotate=31)))

bar.render()